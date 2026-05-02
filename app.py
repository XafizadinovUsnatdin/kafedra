import os
import io
from datetime import datetime
from flask import (Flask, render_template, redirect, url_for, flash,
                   request, send_from_directory, abort, jsonify, make_response)
from flask_login import (LoginManager, login_user, logout_user,
                         login_required, current_user)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from models import db, User, Xodim, IlmiyIsh, UslubiyIsh, Fan, Yangilik, Izoh
from forms import (LoginForm, RegisterForm, IlmiyIshForm, UslubiyIshForm,
                   FanForm, YangilikForm, XodimForm, SettingsForm, IzohForm)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'kafedra-ilmiy-uslubiy-2024-dev-secret')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get(
    'DATABASE_URL', f"sqlite:///{os.path.join(os.path.dirname(__file__), 'instance', 'kafedra.db')}"
)
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

ALLOWED_DOCS   = {'pdf', 'doc', 'docx', 'ppt', 'pptx'}
ALLOWED_IMAGES = {'jpg', 'jpeg', 'png'}

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Iltimos, avval tizimga kiring.'
login_manager.login_message_category = 'warning'


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def allowed_file(filename, allowed):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed


def save_file(file_obj, subfolder='docs'):
    if file_obj and file_obj.filename:
        ext = file_obj.filename.rsplit('.', 1)[-1].lower()
        allowed = ALLOWED_DOCS if ext not in ALLOWED_IMAGES else ALLOWED_IMAGES
        if ext not in allowed:
            return None
        filename = secure_filename(file_obj.filename)
        ts = datetime.utcnow().strftime('%Y%m%d%H%M%S')
        filename = f"{ts}_{filename}"
        dest = os.path.join(app.config['UPLOAD_FOLDER'], subfolder)
        os.makedirs(dest, exist_ok=True)
        file_obj.save(os.path.join(dest, filename))
        return f"{subfolder}/{filename}"
    return None


def admin_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            abort(403)
        return f(*args, **kwargs)
    return decorated


def teacher_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ('admin', 'teacher'):
            flash("Bu sahifaga faqat o'qituvchilar kira oladi.", 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


# ─── AUTH ────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and check_password_hash(user.password_hash, form.password.data):
            if not user.is_active:
                flash('Hisobingiz faollashtirilinmagan.', 'danger')
                return redirect(url_for('login'))
            login_user(user, remember=form.remember.data)
            flash(f'Xush kelibsiz, {user.full_name}!', 'success')
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash("Foydalanuvchi nomi yoki parol noto'g'ri.", 'danger')
    return render_template('auth/login.html', form=form)


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Tizimdan chiqildi.', 'info')
    return redirect(url_for('index'))


@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    form = RegisterForm()
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data).first():
            flash('Bu foydalanuvchi nomi band.', 'danger')
            return render_template('auth/register.html', form=form)
        if User.query.filter_by(email=form.email.data).first():
            flash("Bu email allaqachon ro'yxatdan o'tgan.", 'danger')
            return render_template('auth/register.html', form=form)
        user = User(
            username=form.username.data,
            email=form.email.data,
            full_name=form.full_name.data,
            password_hash=generate_password_hash(form.password.data),
            role=form.role.data,
        )
        db.session.add(user)
        db.session.commit()
        if user.role == 'teacher':
            db.session.add(Xodim(user_id=user.id, kafedra='Umumiy kafedra'))
            db.session.commit()
        flash("Muvaffaqiyatli ro'yxatdan o'tdingiz! Kiring.", 'success')
        return redirect(url_for('login'))
    return render_template('auth/register.html', form=form)


# ─── SETTINGS ────────────────────────────────────────────────────────────────

@app.route('/settings', methods=['GET', 'POST'])
@login_required
def settings():
    form = SettingsForm(obj=current_user)
    if form.validate_on_submit():
        # Email uniqueness check
        existing = User.query.filter_by(email=form.email.data).first()
        if existing and existing.id != current_user.id:
            flash('Bu email boshqa foydalanuvchiga tegishli.', 'danger')
            return render_template('settings.html', form=form)
        current_user.full_name = form.full_name.data
        current_user.email = form.email.data
        if form.new_password.data:
            if not form.current_password.data or \
               not check_password_hash(current_user.password_hash, form.current_password.data):
                flash("Joriy parol noto'g'ri.", 'danger')
                return render_template('settings.html', form=form)
            current_user.password_hash = generate_password_hash(form.new_password.data)
            flash('Parol muvaffaqiyatli yangilandi.', 'success')
        db.session.commit()
        flash('Sozlamalar saqlandi.', 'success')
        return redirect(url_for('settings'))
    return render_template('settings.html', form=form)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    ilmiy_count   = IlmiyIsh.query.filter_by(holat='tasdiqlangan').count()
    uslubiy_count = UslubiyIsh.query.filter_by(holat='tasdiqlangan').count()
    xodim_count   = Xodim.query.count()
    fan_count     = Fan.query.count()
    yangiliklar   = Yangilik.query.filter_by(is_published=True)\
                             .order_by(Yangilik.created_at.desc()).limit(3).all()
    so_nggi_ilmiy   = IlmiyIsh.query.filter_by(holat='tasdiqlangan')\
                               .order_by(IlmiyIsh.created_at.desc()).limit(4).all()
    so_nggi_uslubiy = UslubiyIsh.query.filter_by(holat='tasdiqlangan')\
                                 .order_by(UslubiyIsh.created_at.desc()).limit(4).all()
    return render_template('index.html',
                           ilmiy_count=ilmiy_count, uslubiy_count=uslubiy_count,
                           xodim_count=xodim_count, fan_count=fan_count,
                           yangiliklar=yangiliklar,
                           so_nggi_ilmiy=so_nggi_ilmiy,
                           so_nggi_uslubiy=so_nggi_uslubiy)


@app.route('/dashboard')
@login_required
def dashboard():
    yillar = list(range(2018, datetime.utcnow().year + 1))
    ilmiy_yillik   = [IlmiyIsh.query.filter_by(yil=y, holat='tasdiqlangan').count() for y in yillar]
    uslubiy_yillik = [UslubiyIsh.query.filter_by(yil=y, holat='tasdiqlangan').count() for y in yillar]

    ilmiy_turlar = {
        nom: IlmiyIsh.query.filter_by(tur=val, holat='tasdiqlangan').count()
        for val, nom in [('maqola','Maqola'),('monografiya','Monografiya'),
                         ('patent','Patent'),('dissertatsiya','Dissertatsiya'),('tezis','Tezis')]
    }
    uslubiy_turlar = {
        nom: UslubiyIsh.query.filter_by(tur=val, holat='tasdiqlangan').count()
        for val, nom in [('darslik','Darslik'),("oqquv_qollanma","O'quv qo'llanma"),
                         ("uslubiy_qollanma","Uslubiy qo'llanma"),('dastur','Dastur'),('sillabusi','Sillabusi')]
    }

    my_ilmiy   = IlmiyIsh.query.filter_by(muallif_id=current_user.id).count()
    my_uslubiy = UslubiyIsh.query.filter_by(muallif_id=current_user.id).count()
    my_pending = (IlmiyIsh.query.filter_by(muallif_id=current_user.id, holat='kutilmoqda').count() +
                  UslubiyIsh.query.filter_by(muallif_id=current_user.id, holat='kutilmoqda').count())

    top_ilmiy   = IlmiyIsh.query.filter_by(holat='tasdiqlangan')\
                                 .order_by(IlmiyIsh.views_count.desc()).limit(5).all()
    top_uslubiy = UslubiyIsh.query.filter_by(holat='tasdiqlangan')\
                                   .order_by(UslubiyIsh.downloads_count.desc()).limit(5).all()

    return render_template('dashboard.html',
                           yillar=yillar,
                           ilmiy_yillik=ilmiy_yillik, uslubiy_yillik=uslubiy_yillik,
                           ilmiy_turlar=ilmiy_turlar, uslubiy_turlar=uslubiy_turlar,
                           my_ilmiy=my_ilmiy, my_uslubiy=my_uslubiy, my_pending=my_pending,
                           top_ilmiy=top_ilmiy, top_uslubiy=top_uslubiy)


# ─── ILMIY ISHLAR ─────────────────────────────────────────────────────────────

@app.route('/ilmiy')
def ilmiy_list():
    q     = request.args.get('q', '').strip()
    tur   = request.args.get('tur', '')
    yil   = request.args.get('yil', '')
    autor = request.args.get('autor', '').strip()
    holat = request.args.get('holat', 'tasdiqlangan')
    page  = request.args.get('page', 1, type=int)

    query = IlmiyIsh.query
    if not (current_user.is_authenticated and current_user.role == 'admin'):
        query = query.filter_by(holat='tasdiqlangan')
    elif holat:
        query = query.filter_by(holat=holat)

    if q:
        query = query.filter(
            IlmiyIsh.sarlavha.ilike(f'%{q}%') |
            IlmiyIsh.tavsif.ilike(f'%{q}%') |
            IlmiyIsh.nashr_joyi.ilike(f'%{q}%')
        )
    if tur:
        query = query.filter_by(tur=tur)
    if yil:
        try: query = query.filter_by(yil=int(yil))
        except ValueError: pass
    if autor:
        query = query.join(User).filter(User.full_name.ilike(f'%{autor}%'))

    ishlar = query.order_by(IlmiyIsh.yil.desc(), IlmiyIsh.created_at.desc())\
                  .paginate(page=page, per_page=10)
    yillar = sorted({r[0] for r in db.session.query(IlmiyIsh.yil).distinct()}, reverse=True)
    return render_template('ilmiy/list.html', ishlar=ishlar,
                           q=q, tur=tur, yil=yil, autor=autor, holat=holat, yillar=yillar)


@app.route('/ilmiy/<int:id>')
def ilmiy_detail(id):
    ish = IlmiyIsh.query.get_or_404(id)
    if ish.holat != 'tasdiqlangan':
        if not current_user.is_authenticated:
            abort(404)
        if current_user.role != 'admin' and current_user.id != ish.muallif_id:
            abort(404)
    # view counter
    ish.views_count = (ish.views_count or 0) + 1
    db.session.commit()
    form = IzohForm()
    return render_template('ilmiy/detail.html', ish=ish, form=form)


@app.route('/ilmiy/<int:id>/izoh', methods=['POST'])
@login_required
def ilmiy_izoh(id):
    ish = IlmiyIsh.query.get_or_404(id)
    form = IzohForm()
    if form.validate_on_submit():
        izoh = Izoh(matn=form.matn.data, muallif_id=current_user.id, ilmiy_ish_id=ish.id)
        db.session.add(izoh)
        db.session.commit()
        flash('Izoh qoldirildi.', 'success')
    return redirect(url_for('ilmiy_detail', id=id))


@app.route('/ilmiy/qoshish', methods=['GET', 'POST'])
@login_required
@teacher_required
def ilmiy_qoshish():
    form = IlmiyIshForm()
    if form.validate_on_submit():
        fayl_path = save_file(form.fayl.data)
        ish = IlmiyIsh(
            sarlavha=form.sarlavha.data, muallif_id=current_user.id,
            hammuallif=form.hammuallif.data, tur=form.tur.data,
            yil=form.yil.data, nashr_joyi=form.nashr_joyi.data,
            doi_link=form.doi_link.data, fayl_path=fayl_path,
            tavsif=form.tavsif.data,
            holat='tasdiqlangan' if current_user.role == 'admin' else 'kutilmoqda',
        )
        db.session.add(ish)
        db.session.commit()
        flash("Ilmiy ish qo'shildi. Admin tasdiqlashini kuting.", 'success')
        return redirect(url_for('ilmiy_detail', id=ish.id))
    return render_template('ilmiy/form.html', form=form, title="Ilmiy ish qo'shish")


@app.route('/ilmiy/<int:id>/tahrirlash', methods=['GET', 'POST'])
@login_required
def ilmiy_tahrirlash(id):
    ish = IlmiyIsh.query.get_or_404(id)
    if current_user.role != 'admin' and current_user.id != ish.muallif_id:
        abort(403)
    form = IlmiyIshForm(obj=ish)
    if form.validate_on_submit():
        ish.sarlavha = form.sarlavha.data
        ish.hammuallif = form.hammuallif.data
        ish.tur = form.tur.data
        ish.yil = form.yil.data
        ish.nashr_joyi = form.nashr_joyi.data
        ish.doi_link = form.doi_link.data
        ish.tavsif = form.tavsif.data
        new_file = save_file(form.fayl.data)
        if new_file:
            ish.fayl_path = new_file
        if current_user.role != 'admin':
            ish.holat = 'kutilmoqda'
        db.session.commit()
        flash('Ilmiy ish yangilandi.', 'success')
        return redirect(url_for('ilmiy_detail', id=ish.id))
    return render_template('ilmiy/form.html', form=form, ish=ish, title="Ilmiy ishni tahrirlash")


@app.route('/ilmiy/<int:id>/ochirish', methods=['POST'])
@login_required
def ilmiy_ochirish(id):
    ish = IlmiyIsh.query.get_or_404(id)
    if current_user.role != 'admin' and current_user.id != ish.muallif_id:
        abort(403)
    db.session.delete(ish)
    db.session.commit()
    flash("Ilmiy ish o'chirildi.", 'info')
    return redirect(url_for('ilmiy_list'))


# ─── USLUBIY ISHLAR ───────────────────────────────────────────────────────────

@app.route('/uslubiy')
def uslubiy_list():
    q     = request.args.get('q', '').strip()
    tur   = request.args.get('tur', '')
    yil   = request.args.get('yil', '')
    fan   = request.args.get('fan', '')
    autor = request.args.get('autor', '').strip()
    holat = request.args.get('holat', 'tasdiqlangan')
    page  = request.args.get('page', 1, type=int)

    query = UslubiyIsh.query
    if not (current_user.is_authenticated and current_user.role == 'admin'):
        query = query.filter_by(holat='tasdiqlangan')
    elif holat:
        query = query.filter_by(holat=holat)

    if q:
        query = query.filter(
            UslubiyIsh.sarlavha.ilike(f'%{q}%') |
            UslubiyIsh.tavsif.ilike(f'%{q}%')
        )
    if tur:
        query = query.filter_by(tur=tur)
    if yil:
        try: query = query.filter_by(yil=int(yil))
        except ValueError: pass
    if fan:
        query = query.filter(UslubiyIsh.fan_nomi.ilike(f'%{fan}%'))
    if autor:
        query = query.join(User).filter(User.full_name.ilike(f'%{autor}%'))

    ishlar = query.order_by(UslubiyIsh.yil.desc(), UslubiyIsh.created_at.desc())\
                  .paginate(page=page, per_page=10)
    yillar = sorted({r[0] for r in db.session.query(UslubiyIsh.yil).distinct()}, reverse=True)
    return render_template('uslubiy/list.html', ishlar=ishlar,
                           q=q, tur=tur, yil=yil, fan=fan, autor=autor, holat=holat, yillar=yillar)


@app.route('/uslubiy/<int:id>')
def uslubiy_detail(id):
    ish = UslubiyIsh.query.get_or_404(id)
    if ish.holat != 'tasdiqlangan':
        if not current_user.is_authenticated:
            abort(404)
        if current_user.role != 'admin' and current_user.id != ish.muallif_id:
            abort(404)
    ish.views_count = (ish.views_count or 0) + 1
    db.session.commit()
    form = IzohForm()
    return render_template('uslubiy/detail.html', ish=ish, form=form)


@app.route('/uslubiy/<int:id>/izoh', methods=['POST'])
@login_required
def uslubiy_izoh(id):
    ish = UslubiyIsh.query.get_or_404(id)
    form = IzohForm()
    if form.validate_on_submit():
        izoh = Izoh(matn=form.matn.data, muallif_id=current_user.id, uslubiy_ish_id=ish.id)
        db.session.add(izoh)
        db.session.commit()
        flash('Izoh qoldirildi.', 'success')
    return redirect(url_for('uslubiy_detail', id=id))


@app.route('/uslubiy/qoshish', methods=['GET', 'POST'])
@login_required
@teacher_required
def uslubiy_qoshish():
    form = UslubiyIshForm()
    if form.validate_on_submit():
        fayl_path = save_file(form.fayl.data)
        ish = UslubiyIsh(
            sarlavha=form.sarlavha.data, muallif_id=current_user.id,
            hammuallif=form.hammuallif.data, tur=form.tur.data,
            fan_nomi=form.fan_nomi.data, yil=form.yil.data,
            nashriyot=form.nashriyot.data, fayl_path=fayl_path,
            tavsif=form.tavsif.data,
            holat='tasdiqlangan' if current_user.role == 'admin' else 'kutilmoqda',
        )
        db.session.add(ish)
        db.session.commit()
        flash("Uslubiy ish qo'shildi.", 'success')
        return redirect(url_for('uslubiy_detail', id=ish.id))
    return render_template('uslubiy/form.html', form=form, title="Uslubiy ish qo'shish")


@app.route('/uslubiy/<int:id>/tahrirlash', methods=['GET', 'POST'])
@login_required
def uslubiy_tahrirlash(id):
    ish = UslubiyIsh.query.get_or_404(id)
    if current_user.role != 'admin' and current_user.id != ish.muallif_id:
        abort(403)
    form = UslubiyIshForm(obj=ish)
    if form.validate_on_submit():
        ish.sarlavha = form.sarlavha.data
        ish.hammuallif = form.hammuallif.data
        ish.tur = form.tur.data
        ish.fan_nomi = form.fan_nomi.data
        ish.yil = form.yil.data
        ish.nashriyot = form.nashriyot.data
        ish.tavsif = form.tavsif.data
        new_file = save_file(form.fayl.data)
        if new_file:
            ish.fayl_path = new_file
        if current_user.role != 'admin':
            ish.holat = 'kutilmoqda'
        db.session.commit()
        flash('Uslubiy ish yangilandi.', 'success')
        return redirect(url_for('uslubiy_detail', id=ish.id))
    return render_template('uslubiy/form.html', form=form, ish=ish, title="Uslubiy ishni tahrirlash")


@app.route('/uslubiy/<int:id>/ochirish', methods=['POST'])
@login_required
def uslubiy_ochirish(id):
    ish = UslubiyIsh.query.get_or_404(id)
    if current_user.role != 'admin' and current_user.id != ish.muallif_id:
        abort(403)
    db.session.delete(ish)
    db.session.commit()
    flash("Uslubiy ish o'chirildi.", 'info')
    return redirect(url_for('uslubiy_list'))


# ─── XODIMLAR ─────────────────────────────────────────────────────────────────

@app.route('/xodimlar')
def xodimlar_list():
    xodimlar = (Xodim.query.join(User)
                .filter(User.role == 'teacher', User.is_active == True).all())
    return render_template('xodimlar/list.html', xodimlar=xodimlar)


@app.route('/xodimlar/<int:id>')
def xodim_detail(id):
    xodim   = Xodim.query.get_or_404(id)
    ilmiy   = IlmiyIsh.query.filter_by(muallif_id=xodim.user_id, holat='tasdiqlangan')\
                             .order_by(IlmiyIsh.yil.desc()).all()
    uslubiy = UslubiyIsh.query.filter_by(muallif_id=xodim.user_id, holat='tasdiqlangan')\
                               .order_by(UslubiyIsh.yil.desc()).all()
    return render_template('xodimlar/detail.html', xodim=xodim, ilmiy=ilmiy, uslubiy=uslubiy)


@app.route('/profil/tahrirlash', methods=['GET', 'POST'])
@login_required
def profil_tahrirlash():
    xodim = Xodim.query.filter_by(user_id=current_user.id).first()
    if not xodim:
        xodim = Xodim(user_id=current_user.id)
        db.session.add(xodim)
        db.session.commit()
    form = XodimForm(obj=xodim)
    if form.validate_on_submit():
        xodim.lavozim      = form.lavozim.data
        xodim.ilmiy_daraja = form.ilmiy_daraja.data
        xodim.ilmiy_unvon  = form.ilmiy_unvon.data
        xodim.kafedra      = form.kafedra.data
        xodim.telefon      = form.telefon.data
        xodim.bio          = form.bio.data
        photo = save_file(form.photo.data, subfolder='photos')
        if photo:
            xodim.photo_path = photo
        db.session.commit()
        flash('Profil yangilandi.', 'success')
        return redirect(url_for('xodim_detail', id=xodim.id))
    return render_template('xodimlar/profil_form.html', form=form, xodim=xodim)


# ─── FANLAR ───────────────────────────────────────────────────────────────────

@app.route('/fanlar')
def fanlar_list():
    q       = request.args.get('q', '').strip()
    kurs    = request.args.get('kurs', '')
    semestr = request.args.get('semestr', '')
    query   = Fan.query
    if q:
        query = query.filter(Fan.nomi.ilike(f'%{q}%'))
    if kurs:
        try: query = query.filter_by(kurs=int(kurs))
        except ValueError: pass
    if semestr:
        try: query = query.filter_by(semestr=int(semestr))
        except ValueError: pass
    fanlar = query.order_by(Fan.kurs, Fan.semestr, Fan.nomi).all()
    return render_template('fanlar/list.html', fanlar=fanlar, q=q, kurs=kurs, semestr=semestr)


@app.route('/fanlar/<int:id>')
def fan_detail(id):
    fan     = Fan.query.get_or_404(id)
    uslubiy = UslubiyIsh.query.filter(
        UslubiyIsh.fan_nomi.ilike(f'%{fan.nomi}%'),
        UslubiyIsh.holat == 'tasdiqlangan'
    ).all()
    return render_template('fanlar/detail.html', fan=fan, uslubiy=uslubiy)


@app.route('/fanlar/qoshish', methods=['GET', 'POST'])
@login_required
@teacher_required
def fan_qoshish():
    form = FanForm()
    if form.validate_on_submit():
        fan = Fan(
            nomi=form.nomi.data, kodi=form.kodi.data,
            muallim_id=current_user.id,
            kredit=form.kredit.data, soatlar=form.soatlar.data,
            semestr=form.semestr.data, kurs=form.kurs.data,
            mutaxassislik=form.mutaxassislik.data, tavsif=form.tavsif.data,
        )
        db.session.add(fan)
        db.session.commit()
        flash("Fan qo'shildi.", 'success')
        return redirect(url_for('fan_detail', id=fan.id))
    return render_template('fanlar/form.html', form=form)


# ─── TALABA KABINETI ──────────────────────────────────────────────────────────

@app.route('/talaba/kabinet')
@login_required
def talaba_kabinet():
    fanlar   = Fan.query.order_by(Fan.kurs, Fan.semestr).all()
    uslubiy  = UslubiyIsh.query.filter_by(holat='tasdiqlangan')\
                                .order_by(UslubiyIsh.yil.desc()).limit(20).all()
    ilmiy    = IlmiyIsh.query.filter_by(holat='tasdiqlangan')\
                              .order_by(IlmiyIsh.created_at.desc()).limit(10).all()
    yangiliklar = Yangilik.query.filter_by(is_published=True)\
                                .order_by(Yangilik.created_at.desc()).limit(5).all()
    return render_template('talaba/kabinet.html',
                           fanlar=fanlar, uslubiy=uslubiy,
                           ilmiy=ilmiy, yangiliklar=yangiliklar)


# ─── HISOBOTLAR ───────────────────────────────────────────────────────────────

@app.route('/hisobotlar')
@login_required
def hisobotlar():
    current_year = datetime.utcnow().year
    yil  = request.args.get('yil', current_year, type=int)
    yillar = list(range(2018, current_year + 1))

    xodimlar_stat = []
    for x in Xodim.query.join(User).filter(User.role == 'teacher', User.is_active == True).all():
        il = IlmiyIsh.query.filter_by(muallif_id=x.user_id, yil=yil).count()
        ul = UslubiyIsh.query.filter_by(muallif_id=x.user_id, yil=yil).count()
        xodimlar_stat.append({'xodim': x, 'ilmiy': il, 'uslubiy': ul, 'jami': il + ul})
    xodimlar_stat.sort(key=lambda r: r['jami'], reverse=True)

    return render_template('hisobotlar/index.html',
                           xodimlar_stat=xodimlar_stat, yil=yil, yillar=yillar,
                           ilmiy_jami=IlmiyIsh.query.filter_by(holat='tasdiqlangan').count(),
                           uslubiy_jami=UslubiyIsh.query.filter_by(holat='tasdiqlangan').count(),
                           ilmiy_bu_yil=IlmiyIsh.query.filter_by(holat='tasdiqlangan', yil=yil).count(),
                           uslubiy_bu_yil=UslubiyIsh.query.filter_by(holat='tasdiqlangan', yil=yil).count())


@app.route('/hisobotlar/excel')
@login_required
def hisobotlar_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    yil = request.args.get('yil', datetime.utcnow().year, type=int)

    wb = Workbook()

    # --- Sheet 1: Xodimlar ---
    ws1 = wb.active
    ws1.title = "Xodimlar faolligi"
    header_fill = PatternFill("solid", fgColor="1A5276")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "F.I.Sh.", "Lavozim", "Ilmiy daraja",
               f"Ilmiy ishlar ({yil})", f"Uslubiy ishlar ({yil})", "Jami"]
    ws1.append(headers)
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    xodimlar = Xodim.query.join(User).filter(User.role == 'teacher', User.is_active == True).all()
    rows = []
    for x in xodimlar:
        il = IlmiyIsh.query.filter_by(muallif_id=x.user_id, yil=yil).count()
        ul = UslubiyIsh.query.filter_by(muallif_id=x.user_id, yil=yil).count()
        rows.append((x.user.full_name, x.lavozim or '', x.ilmiy_daraja or '', il, ul, il + ul))
    rows.sort(key=lambda r: r[-1], reverse=True)
    for i, row in enumerate(rows, 1):
        ws1.append([i] + list(row))
        for cell in ws1[i + 1]:
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 18

    # --- Sheet 2: Ilmiy ishlar ---
    ws2 = wb.create_sheet("Ilmiy ishlar")
    ws2.append(["#", "Sarlavha", "Muallif", "Tur", "Yil", "Nashr joyi", "Ko'rishlar"])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for i, ish in enumerate(IlmiyIsh.query.filter_by(holat='tasdiqlangan')
                                          .order_by(IlmiyIsh.yil.desc()).all(), 1):
        ws2.append([i, ish.sarlavha, ish.muallif.full_name,
                    ish.tur, ish.yil, ish.nashr_joyi or '', ish.views_count or 0])
        for cell in ws2[i + 1]:
            cell.border = border
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 25

    # --- Sheet 3: Uslubiy ishlar ---
    ws3 = wb.create_sheet("Uslubiy ishlar")
    ws3.append(["#", "Sarlavha", "Muallif", "Tur", "Fan", "Yil", "Yuklamalar"])
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for i, ish in enumerate(UslubiyIsh.query.filter_by(holat='tasdiqlangan')
                                            .order_by(UslubiyIsh.yil.desc()).all(), 1):
        ws3.append([i, ish.sarlavha, ish.muallif.full_name,
                    ish.tur, ish.fan_nomi or '', ish.yil, ish.downloads_count or 0])
        for cell in ws3[i + 1]:
            cell.border = border
    ws3.column_dimensions['B'].width = 50
    ws3.column_dimensions['C'].width = 25

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = make_response(buf.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=kafedra_hisoboti_{yil}.xlsx'
    return response


@app.route('/hisobotlar/pdf')
@login_required
def hisobotlar_pdf():
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.units import cm

    yil = request.args.get('yil', datetime.utcnow().year, type=int)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                 fontSize=14, spaceAfter=6, alignment=1)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                               fontSize=10, spaceAfter=12, alignment=1)
    head_style = ParagraphStyle('head', parent=styles['Heading2'],
                                fontSize=11, spaceBefore=14, spaceAfter=6)

    header_color = colors.HexColor('#1A5276')
    row_alt_color = colors.HexColor('#EBF5FB')

    story = [
        Paragraph("KAFEDRA ILMIY-USLUBIY ISHLAR HISOBOTI", title_style),
        Paragraph(f"{yil}-yil uchun", sub_style),
        Spacer(1, 0.3*cm),
    ]

    # Umumiy statistika
    story.append(Paragraph("Umumiy ko'rsatkichlar", head_style))
    stats_data = [
        ["Ko'rsatkich", "Miqdor"],
        ["Jami ilmiy ishlar (tasdiqlangan)",
         str(IlmiyIsh.query.filter_by(holat='tasdiqlangan').count())],
        ["Jami uslubiy ishlar (tasdiqlangan)",
         str(UslubiyIsh.query.filter_by(holat='tasdiqlangan').count())],
        [f"{yil}-yil ilmiy ishlar",
         str(IlmiyIsh.query.filter_by(holat='tasdiqlangan', yil=yil).count())],
        [f"{yil}-yil uslubiy ishlar",
         str(UslubiyIsh.query.filter_by(holat='tasdiqlangan', yil=yil).count())],
        ["Kafedra xodimlari soni", str(Xodim.query.count())],
    ]
    t = Table(stats_data, colWidths=[13*cm, 4*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, row_alt_color]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (1,0), (1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
    ]))
    story.append(t)

    # Xodimlar faolligi jadvali
    story.append(Paragraph(f"{yil}-yil xodimlar faolligi jadvali", head_style))
    rows = [["#", "F.I.Sh.", "Lavozim", "Ilmiy", "Uslubiy", "Jami"]]
    xodimlar = Xodim.query.join(User).filter(User.role == 'teacher', User.is_active == True).all()
    data = []
    for x in xodimlar:
        il = IlmiyIsh.query.filter_by(muallif_id=x.user_id, yil=yil).count()
        ul = UslubiyIsh.query.filter_by(muallif_id=x.user_id, yil=yil).count()
        data.append((x.user.full_name, x.lavozim or '', il, ul, il+ul))
    data.sort(key=lambda r: r[-1], reverse=True)
    for i, (name, lavozim, il, ul, jami) in enumerate(data, 1):
        rows.append([str(i), name, lavozim, str(il), str(ul), str(jami)])

    t2 = Table(rows, colWidths=[1*cm, 7*cm, 5*cm, 1.5*cm, 1.5*cm, 1.5*cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, row_alt_color]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (1,1), (2,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(t2)

    doc.build(story)
    buf.seek(0)
    response = make_response(buf.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename=kafedra_hisoboti_{yil}.pdf'
    return response


# ─── ADMIN ────────────────────────────────────────────────────────────────────

@app.route('/admin')
@login_required
@admin_required
def admin_dashboard():
    return render_template('admin/dashboard.html',
                           user_count=User.query.count(),
                           ilmiy_pending=IlmiyIsh.query.filter_by(holat='kutilmoqda').count(),
                           uslubiy_pending=UslubiyIsh.query.filter_by(holat='kutilmoqda').count(),
                           so_nggi_users=User.query.order_by(User.created_at.desc()).limit(5).all())


@app.route('/admin/foydalanuvchilar')
@login_required
@admin_required
def admin_users():
    return render_template('admin/users.html', users=User.query.order_by(User.created_at.desc()).all())


@app.route('/admin/foydalanuvchi/<int:id>/holat', methods=['POST'])
@login_required
@admin_required
def admin_user_holat(id):
    user = User.query.get_or_404(id)
    user.is_active = not user.is_active
    db.session.commit()
    flash(f"{user.full_name} {'faollashtirildi' if user.is_active else 'bloklandi'}.", 'info')
    return redirect(url_for('admin_users'))


@app.route('/admin/kutilmoqda')
@login_required
@admin_required
def admin_kutilmoqda():
    return render_template('admin/kutilmoqda.html',
                           ilmiy=IlmiyIsh.query.filter_by(holat='kutilmoqda')
                                               .order_by(IlmiyIsh.created_at.desc()).all(),
                           uslubiy=UslubiyIsh.query.filter_by(holat='kutilmoqda')
                                                   .order_by(UslubiyIsh.created_at.desc()).all())


@app.route('/admin/tasdiqlash/ilmiy/<int:id>', methods=['POST'])
@login_required
@admin_required
def admin_ilmiy_tasdiqlash(id):
    ish = IlmiyIsh.query.get_or_404(id)
    ish.holat = request.form.get('holat', 'tasdiqlangan')
    db.session.commit()
    flash(f'Ilmiy ish {ish.holat}.', 'success')
    return redirect(url_for('admin_kutilmoqda'))


@app.route('/admin/tasdiqlash/uslubiy/<int:id>', methods=['POST'])
@login_required
@admin_required
def admin_uslubiy_tasdiqlash(id):
    ish = UslubiyIsh.query.get_or_404(id)
    ish.holat = request.form.get('holat', 'tasdiqlangan')
    db.session.commit()
    flash(f'Uslubiy ish {ish.holat}.', 'success')
    return redirect(url_for('admin_kutilmoqda'))


# ─── YANGILIKLAR ──────────────────────────────────────────────────────────────

@app.route('/yangiliklar')
def yangiliklar_list():
    yangiliklar = Yangilik.query.filter_by(is_published=True)\
                                .order_by(Yangilik.created_at.desc()).all()
    return render_template('yangiliklar.html', yangiliklar=yangiliklar)


@app.route('/yangiliklar/<int:id>')
def yangilik_detail(id):
    return render_template('yangilik_detail.html', yangilik=Yangilik.query.get_or_404(id))


@app.route('/yangiliklar/qoshish', methods=['GET', 'POST'])
@login_required
@admin_required
def yangilik_qoshish():
    form = YangilikForm()
    if form.validate_on_submit():
        db.session.add(Yangilik(sarlavha=form.sarlavha.data, matn=form.matn.data,
                                muallif_id=current_user.id, is_published=form.is_published.data))
        db.session.commit()
        flash("Yangilik qo'shildi.", 'success')
        return redirect(url_for('yangiliklar_list'))
    return render_template('yangilik_form.html', form=form)


# ─── FILE DOWNLOAD ────────────────────────────────────────────────────────────

@app.route('/uploads/<path:filename>')
@login_required
def download_file(filename):
    # Track downloads for docs (not photos)
    if filename.startswith('docs/'):
        fname = filename[5:]
        for ish in list(IlmiyIsh.query.filter_by(fayl_path=filename).all()) + \
                   list(UslubiyIsh.query.filter_by(fayl_path=filename).all()):
            ish.downloads_count = (ish.downloads_count or 0) + 1
        db.session.commit()
        return send_from_directory(
            os.path.join(app.config['UPLOAD_FOLDER'], 'docs'), fname, as_attachment=True)
    if filename.startswith('photos/'):
        fname = filename[7:]
        return send_from_directory(
            os.path.join(app.config['UPLOAD_FOLDER'], 'photos'), fname)
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)


# ─── API ──────────────────────────────────────────────────────────────────────

@app.route('/api/statistika')
def api_statistika():
    yillar = list(range(2018, datetime.utcnow().year + 1))
    return jsonify({
        'yillar': yillar,
        'ilmiy':  [IlmiyIsh.query.filter_by(yil=y, holat='tasdiqlangan').count() for y in yillar],
        'uslubiy':[UslubiyIsh.query.filter_by(yil=y, holat='tasdiqlangan').count() for y in yillar],
        'jami_ilmiy':   IlmiyIsh.query.filter_by(holat='tasdiqlangan').count(),
        'jami_uslubiy': UslubiyIsh.query.filter_by(holat='tasdiqlangan').count(),
        'xodimlar':     Xodim.query.count(),
        'fanlar':       Fan.query.count(),
    })


# ─── ERROR HANDLERS ───────────────────────────────────────────────────────────

@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403


@app.errorhandler(404)
def not_found(e):
    return render_template('errors/404.html'), 404


@app.context_processor
def inject_globals():
    pending_count = 0
    if current_user.is_authenticated and current_user.role == 'admin':
        pending_count = (IlmiyIsh.query.filter_by(holat='kutilmoqda').count() +
                         UslubiyIsh.query.filter_by(holat='kutilmoqda').count())
    return dict(pending_count=pending_count, now=datetime.utcnow())


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, port=5000)
